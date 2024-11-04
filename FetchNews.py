import requests
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import time
import os
import argparse

# Function to fetch news article titles in English based on the search term and API key
def fetch_latest_news_titles(search_term, api_key):
    url = f'https://newsapi.org/v2/everything?q={search_term}&language=en&sortBy=publishedAt&pageSize=30&apiKey={api_key}'
    response = requests.get(url)
    articles = response.json().get("articles", [])
    
    # Filter articles to include only those where the search term appears in the title or description
    return [
        {
            "title": article["title"],
            "link": article["url"],
            "published_date": article["publishedAt"]
        }
        for article in articles
        if search_term.lower() in (article["title"] or "").lower() or search_term.lower() in (article["description"] or "").lower()
    ]

# Retry mechanism for updating Excel in case of PermissionError
def update_excel_with_titles(titles):
    workbook_path = 'AI_Latest_Updates.xlsx'
    max_retries = 5  # Maximum retry attempts
    retry_delay = 2  # Seconds to wait between retries

    # Check if the file exists; if not, create it with headers
    if not os.path.exists(workbook_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "AI_Latest_Updates"
        sheet.append(["Post", "Article date", "Link"])  # Updated header to "Article date"
        workbook.save(workbook_path)
        print(f"Created new file with headers: {workbook_path}")

    for attempt in range(max_retries):
        try:
            # Load the workbook and select the active sheet
            workbook = openpyxl.load_workbook(workbook_path)
            sheet = workbook.active

            # Find the next available row in the sheet
            next_row = sheet.max_row + 1

            # Update the sheet with new titles, links, and formatted publication date, only if the title is not a duplicate
            for title_data in titles:
                if not title_exists(sheet, title_data['title']):  # Check for duplicates
                    # Format the published date to show only the date part
                    article_date = datetime.strptime(title_data['published_date'], "%Y-%m-%dT%H:%M:%SZ").date()
                    sheet[f"A{next_row}"].value = title_data['title']
                    sheet[f"B{next_row}"].value = article_date  # Add the formatted date
                    sheet[f"C{next_row}"].value = title_data['link']
                    next_row += 1
                else:
                    print(f"Skipping duplicate title: {title_data['title']}")

            # Save the workbook
            workbook.save(workbook_path)
            print("Excel sheet updated successfully with article titles.")
            break  # Exit loop if successful

        except PermissionError:
            print(f"Permission denied for {workbook_path}. Retrying in {retry_delay} seconds...")
            time.sleep(retry_delay)
        
        except Exception as e:
            print(f"An error occurred: {e}")
            break
    else:
        print("Failed to access the Excel file after multiple attempts. Please close the file if it’s open and try again.")

def title_exists(sheet, title):
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):  # Assume titles are in column A starting from row 2
        if row[0] == title:
            return True
    return False

def main(search_term, api_key):
    # Step 1: Fetch the latest news titles in English for the given search term and API key
    titles = fetch_latest_news_titles(search_term, api_key)
    
    # Step 2: Update the Excel sheet with titles, links, and publication date
    update_excel_with_titles(titles)

if __name__ == "__main__":
    # Parse the command-line arguments for the search term and API key
    parser = argparse.ArgumentParser(description="Fetch and save latest news articles based on a search term and API key.")
    parser.add_argument("search_term", type=str, help="The search term to look for in news articles (e.g., 'soccer').")
    parser.add_argument("api_key", type=str, help="Your News API key for accessing the News API.")
    args = parser.parse_args()
    
    # Run the main function with the provided search term and API key
    main(args.search_term, args.api_key)
